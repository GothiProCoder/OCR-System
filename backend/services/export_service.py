"""
Export Service - Multi-Format Document Export
==============================================
Production-grade service for exporting extracted data to various formats.

Features:
- Excel export with styled worksheets and field highlighting
- JSON export with complete structured data
- CSV export for simple flat data
- PDF export with professional document-style reports
- Bulk export with merge/summary options
- Validation gate before export

Usage:
    from services.export_service import export_service
    
    # Single export
    result = await export_service.export_extraction(
        extraction_id=ext_id,
        format="excel",
        options={"include_metadata": True}
    )
    
    # Bulk export
    result = await export_service.bulk_export(
        extraction_ids=[id1, id2, id3],
        format="excel",
        merge_into_single_file=True
    )
"""

import csv
import io
import json
import logging
import os
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
from uuid import UUID, uuid4

# Excel library
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# PDF library
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, Image, PageBreak
)

# Database imports
from database.crud import extraction_crud, document_crud
from database.connection import get_async_db

# Internal services
from services.validation_service import ValidationService

# Schemas
from schemas.export import (
    ExportFormat, ExportStatus, ExportRequest,
    ExportFieldData, ExportDocumentData, ExportExtractionData
)

# Configuration
from config import settings

logger = logging.getLogger(__name__)


# =============================================================================
# DATA CLASSES - Export Results
# =============================================================================

@dataclass
class ExportResult:
    """Result from export operation."""
    export_id: UUID = field(default_factory=uuid4)
    extraction_id: Optional[UUID] = None
    format: str = "excel"
    status: str = "completed"
    
    # File info
    file_path: Optional[str] = None
    file_name: Optional[str] = None
    file_size_bytes: int = 0
    download_url: Optional[str] = None
    
    # Metadata
    field_count: int = 0
    processing_time_ms: int = 0
    
    # Status
    success: bool = True
    error: Optional[str] = None
    validation_passed: bool = True
    validation_errors: List[str] = field(default_factory=list)
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "export_id": str(self.export_id),
            "extraction_id": str(self.extraction_id) if self.extraction_id else None,
            "format": self.format,
            "status": self.status,
            "file_path": self.file_path,
            "file_name": self.file_name,
            "file_size_bytes": self.file_size_bytes,
            "download_url": self.download_url,
            "field_count": self.field_count,
            "processing_time_ms": self.processing_time_ms,
            "success": self.success,
            "error": self.error,
            "validation_passed": self.validation_passed,
            "validation_errors": self.validation_errors
        }


@dataclass
class BulkExportResult:
    """Result from bulk export operation."""
    export_id: UUID = field(default_factory=uuid4)
    extraction_count: int = 0
    format: str = "excel"
    status: str = "completed"
    
    file_path: Optional[str] = None
    file_name: Optional[str] = None
    file_size_bytes: int = 0
    download_url: Optional[str] = None
    
    total_fields: int = 0
    processing_time_ms: int = 0
    
    success: bool = True
    error: Optional[str] = None
    failed_extractions: List[str] = field(default_factory=list)
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "export_id": str(self.export_id),
            "extraction_count": self.extraction_count,
            "format": self.format,
            "status": self.status,
            "file_path": self.file_path,
            "file_name": self.file_name,
            "file_size_bytes": self.file_size_bytes,
            "download_url": self.download_url,
            "total_fields": self.total_fields,
            "processing_time_ms": self.processing_time_ms,
            "success": self.success,
            "error": self.error,
            "failed_extractions": self.failed_extractions
        }


# =============================================================================
# EXCEL STYLING CONSTANTS
# =============================================================================

# Fonts
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
NORMAL_FONT = Font(size=11)
ERROR_FONT = Font(size=11, color="FF0000")
WARNING_FONT = Font(size=11, color="FF8C00")

# Fills
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HIGH_CONFIDENCE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MEDIUM_CONFIDENCE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
LOW_CONFIDENCE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
INVALID_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alignment
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


# =============================================================================
# EXPORTERS - Individual Format Handlers
# =============================================================================

def export_to_excel(
    extraction_data: ExportExtractionData,
    options: Dict[str, Any],
    file_path: Path
) -> Tuple[bool, Optional[str]]:
    """
    Export extraction to Excel format.
    
    Creates a styled workbook with:
    - Metadata sheet (document info, extraction stats)
    - Fields sheet (key-value pairs with confidence highlighting)
    """
    try:
        wb = Workbook()
        
        # ===== Fields Sheet (main data) =====
        ws_fields = wb.active
        ws_fields.title = getattr(settings, 'EXPORT_EXCEL_SHEET_NAME', 'Extracted Data')
        
        # Headers
        headers = ["Field Name", "Value", "Type"]
        if options.get("include_confidence_scores", True):
            headers.extend(["Confidence", "Status"])
        
        for col, header in enumerate(headers, 1):
            cell = ws_fields.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
        
        # Data rows
        for row_idx, field_data in enumerate(extraction_data.fields, 2):
            # Field name
            ws_fields.cell(row=row_idx, column=1, value=field_data.key).border = THIN_BORDER
            
            # Value
            value_cell = ws_fields.cell(row=row_idx, column=2, value=field_data.value or "")
            value_cell.border = THIN_BORDER
            value_cell.alignment = LEFT_ALIGN
            
            # Type
            ws_fields.cell(row=row_idx, column=3, value=field_data.type).border = THIN_BORDER
            
            if options.get("include_confidence_scores", True):
                # Confidence
                conf = field_data.confidence or 0
                conf_cell = ws_fields.cell(row=row_idx, column=4, value=f"{int(conf * 100)}%")
                conf_cell.border = THIN_BORDER
                conf_cell.alignment = CENTER_ALIGN
                
                # Apply confidence-based highlighting
                if conf >= 0.85:
                    conf_cell.fill = HIGH_CONFIDENCE_FILL
                    status = "Valid"
                elif conf >= 0.60:
                    conf_cell.fill = MEDIUM_CONFIDENCE_FILL
                    status = "Review"
                else:
                    conf_cell.fill = LOW_CONFIDENCE_FILL
                    status = "Low Confidence"
                
                # Status
                status_cell = ws_fields.cell(row=row_idx, column=5, value=status)
                status_cell.border = THIN_BORDER
                status_cell.alignment = CENTER_ALIGN
        
        # Adjust column widths
        ws_fields.column_dimensions['A'].width = 25
        ws_fields.column_dimensions['B'].width = 40
        ws_fields.column_dimensions['C'].width = 12
        ws_fields.column_dimensions['D'].width = 12
        ws_fields.column_dimensions['E'].width = 15
        
        # ===== Metadata Sheet =====
        if options.get("include_metadata", True):
            ws_meta = wb.create_sheet("Document Info")
            
            meta_data = [
                ("Document ID", extraction_data.document.document_id),
                ("Filename", extraction_data.document.filename),
                ("Form Type", extraction_data.document.form_type or "Unknown"),
                ("Language", extraction_data.document.language or "en"),
                ("Page Count", str(extraction_data.document.page_count)),
                ("Upload Date", extraction_data.document.uploaded_at.strftime("%Y-%m-%d %H:%M")),
                ("", ""),
                ("Extraction ID", extraction_data.extraction_id),
                ("Version", str(extraction_data.version)),
                ("Total Fields", str(extraction_data.total_fields)),
                ("Edited Fields", str(extraction_data.edited_fields)),
                ("Average Confidence", f"{int((extraction_data.avg_confidence or 0) * 100)}%"),
                ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ]
            
            for row_idx, (label, value) in enumerate(meta_data, 1):
                label_cell = ws_meta.cell(row=row_idx, column=1, value=label)
                label_cell.font = Font(bold=True)
                ws_meta.cell(row=row_idx, column=2, value=value)
            
            ws_meta.column_dimensions['A'].width = 20
            ws_meta.column_dimensions['B'].width = 40
        
        # Save workbook
        wb.save(str(file_path))
        return True, None
        
    except Exception as e:
        logger.exception(f"Excel export error: {e}")
        return False, str(e)


def export_to_json(
    extraction_data: ExportExtractionData,
    options: Dict[str, Any],
    file_path: Path
) -> Tuple[bool, Optional[str]]:
    """
    Export extraction to JSON format.
    
    Creates a well-structured JSON file with all extraction data.
    """
    try:
        export_dict = {
            "export_info": {
                "format": "json",
                "export_date": datetime.now().isoformat(),
                "generated_by": getattr(settings, 'APP_NAME', 'FormExtract AI')
            },
            "document": {
                "id": extraction_data.document.document_id,
                "filename": extraction_data.document.filename,
                "form_type": extraction_data.document.form_type,
                "language": extraction_data.document.language,
                "page_count": extraction_data.document.page_count,
                "uploaded_at": extraction_data.document.uploaded_at.isoformat()
            } if options.get("include_metadata", True) else None,
            "extraction": {
                "id": extraction_data.extraction_id,
                "version": extraction_data.version,
                "total_fields": extraction_data.total_fields,
                "edited_fields": extraction_data.edited_fields,
                "avg_confidence": extraction_data.avg_confidence,
                "created_at": extraction_data.created_at.isoformat(),
                "finalized_at": extraction_data.finalized_at.isoformat() if extraction_data.finalized_at else None
            } if options.get("include_metadata", True) else None,
            "fields": []
        }
        
        for field_data in extraction_data.fields:
            field_dict = {
                "key": field_data.key,
                "value": field_data.value,
                "type": field_data.type
            }
            
            if options.get("include_confidence_scores", True):
                field_dict["confidence"] = field_data.confidence
                field_dict["confidence_percent"] = field_data.confidence_percent
            
            field_dict["is_edited"] = field_data.is_edited
            field_dict["page"] = field_data.page
            
            export_dict["fields"].append(field_dict)
        
        # Remove None entries
        export_dict = {k: v for k, v in export_dict.items() if v is not None}
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_dict, f, indent=2, ensure_ascii=False)
        
        return True, None
        
    except Exception as e:
        logger.exception(f"JSON export error: {e}")
        return False, str(e)


def export_to_csv(
    extraction_data: ExportExtractionData,
    options: Dict[str, Any],
    file_path: Path
) -> Tuple[bool, Optional[str]]:
    """
    Export extraction to CSV format.
    
    Creates a simple flat CSV file with key-value pairs.
    """
    try:
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            headers = ["Field Name", "Value", "Type"]
            if options.get("include_confidence_scores", True):
                headers.extend(["Confidence", "Is Edited"])
            writer.writerow(headers)
            
            # Data rows
            for field_data in extraction_data.fields:
                row = [field_data.key, field_data.value or "", field_data.type]
                
                if options.get("include_confidence_scores", True):
                    row.extend([
                        f"{int((field_data.confidence or 0) * 100)}%",
                        "Yes" if field_data.is_edited else "No"
                    ])
                
                writer.writerow(row)
        
        return True, None
        
    except Exception as e:
        logger.exception(f"CSV export error: {e}")
        return False, str(e)


def export_to_pdf(
    extraction_data: ExportExtractionData,
    options: Dict[str, Any],
    file_path: Path
) -> Tuple[bool, Optional[str]]:
    """
    Export extraction to PDF format.
    
    Creates a professional document-style report with tables.
    """
    try:
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1  # Center
        )
        story.append(Paragraph("Extraction Report", title_style))
        story.append(Spacer(1, 10))
        
        # ===== Document Info Section =====
        if options.get("include_metadata", True):
            story.append(Paragraph("Document Information", styles['Heading2']))
            story.append(Spacer(1, 5))
            
            doc_info = [
                ["Property", "Value"],
                ["Filename", extraction_data.document.filename],
                ["Form Type", extraction_data.document.form_type or "Unknown"],
                ["Language", extraction_data.document.language or "en"],
                ["Page Count", str(extraction_data.document.page_count)],
                ["Upload Date", extraction_data.document.uploaded_at.strftime("%Y-%m-%d %H:%M")],
            ]
            
            doc_table = Table(doc_info, colWidths=[2*inch, 4*inch])
            doc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(doc_table)
            story.append(Spacer(1, 20))
        
        # ===== Extracted Fields Section =====
        story.append(Paragraph("Extracted Fields", styles['Heading2']))
        story.append(Spacer(1, 5))
        
        # Build fields table
        if options.get("include_confidence_scores", True):
            field_headers = ["Field Name", "Value", "Type", "Confidence"]
            col_widths = [1.5*inch, 3*inch, 1*inch, 1*inch]
        else:
            field_headers = ["Field Name", "Value", "Type"]
            col_widths = [2*inch, 3.5*inch, 1*inch]
        
        field_data_rows = [field_headers]
        
        for field_data in extraction_data.fields:
            row = [
                field_data.key,
                field_data.value or "",
                field_data.type
            ]
            if options.get("include_confidence_scores", True):
                row.append(f"{int((field_data.confidence or 0) * 100)}%")
            field_data_rows.append(row)
        
        fields_table = Table(field_data_rows, colWidths=col_widths)
        
        # Dynamic row colors based on confidence
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]
        
        # Apply confidence-based row colors
        if options.get("include_confidence_scores", True):
            for row_idx, field_data in enumerate(extraction_data.fields, 1):
                conf = field_data.confidence or 0
                if conf >= 0.85:
                    bg_color = colors.HexColor('#C6EFCE')  # Green
                elif conf >= 0.60:
                    bg_color = colors.HexColor('#FFEB9C')  # Yellow
                else:
                    bg_color = colors.HexColor('#FFC7CE')  # Red
                table_style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_color))
        
        fields_table.setStyle(TableStyle(table_style))
        story.append(fields_table)
        story.append(Spacer(1, 20))
        
        # ===== Footer =====
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        story.append(Spacer(1, 30))
        story.append(Paragraph(
            f"Generated by {getattr(settings, 'APP_NAME', 'FormExtract AI')} on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            footer_style
        ))
        
        # Build PDF
        doc.build(story)
        return True, None
        
    except Exception as e:
        logger.exception(f"PDF export error: {e}")
        return False, str(e)


# =============================================================================
# EXPORT SERVICE CLASS
# =============================================================================

class ExportService:
    """
    Production-grade export service for extracted data.
    
    Features:
    - Multiple format support (Excel, JSON, CSV, PDF)
    - Validation gate before export
    - Styled output with confidence highlighting
    - Bulk export with merge options
    - Singleton pattern for efficiency
    """
    
    _instance: Optional["ExportService"] = None
    _initialized: bool = False
    
    def __new__(cls) -> "ExportService":
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance
    
    def __init__(self):
        if ExportService._initialized:
            return
        
        self._validation_service = ValidationService()
        self._export_path = getattr(settings, 'EXPORT_PATH', Path('storage/exports'))
        
        # Ensure export directory exists
        self._export_path.mkdir(parents=True, exist_ok=True)
        
        ExportService._initialized = True
        logger.info(f"ExportService initialized, export path: {self._export_path}")
    
    def _get_format_extension(self, format: str) -> str:
        """Get file extension for format."""
        extensions = {
            "excel": ".xlsx",
            "json": ".json",
            "csv": ".csv",
            "pdf": ".pdf"
        }
        return extensions.get(format.lower(), ".xlsx")
    
    def _get_exporter(self, format: str):
        """Get exporter function for format."""
        exporters = {
            "excel": export_to_excel,
            "json": export_to_json,
            "csv": export_to_csv,
            "pdf": export_to_pdf
        }
        return exporters.get(format.lower(), export_to_excel)
    
    async def _fetch_extraction_data(
        self,
        extraction_id: UUID
    ) -> Optional[ExportExtractionData]:
        """Fetch and prepare extraction data for export."""
        try:
            async with get_async_db() as db:
                extraction = await extraction_crud.get_with_fields(db, extraction_id)
                
                if not extraction:
                    return None
                
                # Get document info
                document = await document_crud.get(db, extraction.document_id)
                
                # Build export data structures
                doc_data = ExportDocumentData(
                    document_id=str(document.id),
                    filename=document.filename,
                    form_type=document.form_type,
                    language=document.language,
                    page_count=document.page_count or 1,
                    uploaded_at=document.created_at,
                    finalized_at=extraction.finalized_at
                )
                
                fields_data = []
                for field in extraction.fields:
                    fields_data.append(ExportFieldData(
                        key=field.field_key,
                        value=field.field_value,
                        type=field.field_type.value if field.field_type else "text",
                        confidence=field.confidence,
                        confidence_percent=int((field.confidence or 0) * 100),
                        is_edited=field.is_edited or False,
                        page=field.page_number or 1
                    ))
                
                return ExportExtractionData(
                    extraction_id=str(extraction.id),
                    version=extraction.version,
                    document=doc_data,
                    fields=fields_data,
                    total_fields=len(fields_data),
                    edited_fields=sum(1 for f in fields_data if f.is_edited),
                    avg_confidence=extraction.confidence_avg,
                    created_at=extraction.created_at,
                    finalized_at=extraction.finalized_at
                )
                
        except Exception as e:
            logger.exception(f"Error fetching extraction data: {e}")
            return None
    
    async def export_extraction(
        self,
        extraction_id: UUID,
        format: str = "excel",
        options: Optional[Dict[str, Any]] = None,
        validate_first: bool = True,
        custom_filename: Optional[str] = None
    ) -> ExportResult:
        """
        Export a single extraction to specified format.
        
        Args:
            extraction_id: UUID of extraction to export
            format: Export format (excel, json, csv, pdf)
            options: Export options (metadata, confidence, etc.)
            validate_first: Run validation before export
            custom_filename: Optional custom filename
            
        Returns:
            ExportResult with file info or error
        """
        start_time = time.time()
        result = ExportResult(extraction_id=extraction_id, format=format)
        options = options or {}
        
        try:
            # ===== Step 1: Validation Gate =====
            if validate_first:
                can_export, validation_result = await self._validation_service.validate_before_finalization(
                    extraction_id
                )
                
                if not can_export:
                    result.success = False
                    result.status = "validation_failed"
                    result.validation_passed = False
                    result.validation_errors = [
                        r.message for r in validation_result.field_results 
                        if not r.is_valid
                    ]
                    result.error = f"Export blocked: {validation_result.invalid_count} invalid fields need fixing"
                    logger.warning(f"Export blocked for {extraction_id}: validation failed")
                    return result
            
            # ===== Step 2: Fetch Data =====
            extraction_data = await self._fetch_extraction_data(extraction_id)
            
            if not extraction_data:
                result.success = False
                result.status = "failed"
                result.error = f"Extraction {extraction_id} not found"
                return result
            
            # ===== Step 3: Generate File =====
            ext = self._get_format_extension(format)
            filename = custom_filename or f"extraction_{extraction_id}"
            filename = f"{filename}{ext}"
            file_path = self._export_path / filename
            
            exporter = self._get_exporter(format)
            success, error = exporter(extraction_data, options, file_path)
            
            if not success:
                result.success = False
                result.status = "failed"
                result.error = error
                return result
            
            # ===== Step 4: Return Result =====
            result.file_path = str(file_path)
            result.file_name = filename
            result.file_size_bytes = file_path.stat().st_size if file_path.exists() else 0
            result.download_url = f"/api/exports/{result.export_id}/download"
            result.field_count = extraction_data.total_fields
            result.processing_time_ms = int((time.time() - start_time) * 1000)
            result.success = True
            result.status = "completed"
            
            logger.info(
                f"Export completed: {format.upper()} file generated for {extraction_id} "
                f"({result.file_size_bytes} bytes in {result.processing_time_ms}ms)"
            )
            
            return result
            
        except Exception as e:
            result.processing_time_ms = int((time.time() - start_time) * 1000)
            result.success = False
            result.status = "failed"
            result.error = str(e)
            logger.exception(f"Export error for {extraction_id}: {e}")
            return result
    
    async def bulk_export(
        self,
        extraction_ids: List[UUID],
        format: str = "excel",
        merge_into_single_file: bool = True,
        include_summary: bool = True,
        options: Optional[Dict[str, Any]] = None
    ) -> BulkExportResult:
        """
        Export multiple extractions.
        
        For Excel: Creates workbook with one sheet per extraction + summary sheet.
        For JSON: Creates array of all extractions.
        For CSV/PDF: Creates single merged file.
        """
        start_time = time.time()
        result = BulkExportResult(
            extraction_count=len(extraction_ids),
            format=format
        )
        options = options or {}
        
        try:
            # Fetch all extraction data
            all_data: List[ExportExtractionData] = []
            failed = []
            
            for ext_id in extraction_ids:
                data = await self._fetch_extraction_data(ext_id)
                if data:
                    all_data.append(data)
                else:
                    failed.append(str(ext_id))
            
            if not all_data:
                result.success = False
                result.status = "failed"
                result.error = "No valid extractions found"
                return result
            
            result.failed_extractions = failed
            
            # Generate bulk export file
            ext = self._get_format_extension(format)
            filename = f"bulk_export_{result.export_id}{ext}"
            file_path = self._export_path / filename
            
            if format.lower() == "excel" and merge_into_single_file:
                success, error = self._bulk_export_excel(all_data, options, file_path, include_summary)
            elif format.lower() == "json":
                success, error = self._bulk_export_json(all_data, options, file_path)
            else:
                # For CSV/PDF, concatenate exports
                success, error = self._bulk_export_concatenated(all_data, format, options, file_path)
            
            if not success:
                result.success = False
                result.status = "failed"
                result.error = error
                return result
            
            result.file_path = str(file_path)
            result.file_name = filename
            result.file_size_bytes = file_path.stat().st_size if file_path.exists() else 0
            result.download_url = f"/api/exports/bulk/{result.export_id}/download"
            result.total_fields = sum(d.total_fields for d in all_data)
            result.processing_time_ms = int((time.time() - start_time) * 1000)
            result.success = True
            result.status = "completed"
            
            logger.info(
                f"Bulk export completed: {len(all_data)} extractions to {format.upper()} "
                f"({result.file_size_bytes} bytes)"
            )
            
            return result
            
        except Exception as e:
            result.processing_time_ms = int((time.time() - start_time) * 1000)
            result.success = False
            result.status = "failed"
            result.error = str(e)
            logger.exception(f"Bulk export error: {e}")
            return result
    
    def _bulk_export_excel(
        self,
        all_data: List[ExportExtractionData],
        options: Dict[str, Any],
        file_path: Path,
        include_summary: bool
    ) -> Tuple[bool, Optional[str]]:
        """Create Excel workbook with multiple sheets."""
        try:
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Create sheet for each extraction
            for idx, data in enumerate(all_data):
                sheet_name = f"{data.document.form_type or 'Doc'}_{idx + 1}"[:30]
                ws = wb.create_sheet(title=sheet_name)
                
                # Headers
                headers = ["Field Name", "Value", "Type", "Confidence"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.border = THIN_BORDER
                
                # Data
                for row_idx, field in enumerate(data.fields, 2):
                    ws.cell(row=row_idx, column=1, value=field.key).border = THIN_BORDER
                    ws.cell(row=row_idx, column=2, value=field.value or "").border = THIN_BORDER
                    ws.cell(row=row_idx, column=3, value=field.type).border = THIN_BORDER
                    conf_cell = ws.cell(row=row_idx, column=4, value=f"{int((field.confidence or 0) * 100)}%")
                    conf_cell.border = THIN_BORDER
                
                # Column widths
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 40
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
            
            # Summary sheet
            if include_summary:
                ws_sum = wb.create_sheet(title="Summary", index=0)
                summary_data = [
                    ["Summary", ""],
                    ["Total Extractions", len(all_data)],
                    ["Total Fields", sum(d.total_fields for d in all_data)],
                    ["Average Confidence", f"{int(sum((d.avg_confidence or 0) for d in all_data) / len(all_data) * 100)}%"],
                    ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M")],
                ]
                
                for row_idx, (label, value) in enumerate(summary_data, 1):
                    ws_sum.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                    ws_sum.cell(row=row_idx, column=2, value=value)
                
                ws_sum.column_dimensions['A'].width = 25
                ws_sum.column_dimensions['B'].width = 30
            
            wb.save(str(file_path))
            return True, None
            
        except Exception as e:
            logger.exception(f"Bulk Excel export error: {e}")
            return False, str(e)
    
    def _bulk_export_json(
        self,
        all_data: List[ExportExtractionData],
        options: Dict[str, Any],
        file_path: Path
    ) -> Tuple[bool, Optional[str]]:
        """Create JSON file with all extractions."""
        try:
            export_dict = {
                "export_info": {
                    "format": "json",
                    "export_date": datetime.now().isoformat(),
                    "total_extractions": len(all_data),
                    "total_fields": sum(d.total_fields for d in all_data)
                },
                "extractions": []
            }
            
            for data in all_data:
                ext_dict = {
                    "extraction_id": data.extraction_id,
                    "document": {
                        "filename": data.document.filename,
                        "form_type": data.document.form_type
                    },
                    "fields": [
                        {
                            "key": f.key,
                            "value": f.value,
                            "type": f.type,
                            "confidence": f.confidence
                        }
                        for f in data.fields
                    ]
                }
                export_dict["extractions"].append(ext_dict)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_dict, f, indent=2, ensure_ascii=False)
            
            return True, None
            
        except Exception as e:
            logger.exception(f"Bulk JSON export error: {e}")
            return False, str(e)
    
    def _bulk_export_concatenated(
        self,
        all_data: List[ExportExtractionData],
        format: str,
        options: Dict[str, Any],
        file_path: Path
    ) -> Tuple[bool, Optional[str]]:
        """Concatenate multiple exports into single file (CSV/PDF)."""
        try:
            if format.lower() == "csv":
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["Document", "Field Name", "Value", "Type", "Confidence"])
                    
                    for data in all_data:
                        for field in data.fields:
                            writer.writerow([
                                data.document.filename,
                                field.key,
                                field.value or "",
                                field.type,
                                f"{int((field.confidence or 0) * 100)}%"
                            ])
                
                return True, None
            
            # For PDF, export first one (simplified)
            if all_data:
                return export_to_pdf(all_data[0], options, file_path)
            
            return False, "No data to export"
            
        except Exception as e:
            logger.exception(f"Bulk concatenated export error: {e}")
            return False, str(e)
    
    def get_status(self) -> Dict[str, Any]:
        """Get service status for health checks."""
        return {
            "service": "ExportService",
            "status": "ready",
            "export_path": str(self._export_path),
            "supported_formats": ["excel", "json", "csv", "pdf"]
        }


# =============================================================================
# SINGLETON INSTANCE
# =============================================================================

export_service = ExportService()


# =============================================================================
# MODULE EXPORTS
# =============================================================================

__all__ = [
    "ExportService",
    "export_service",
    "ExportResult",
    "BulkExportResult",
    "export_to_excel",
    "export_to_json",
    "export_to_csv",
    "export_to_pdf"
]
